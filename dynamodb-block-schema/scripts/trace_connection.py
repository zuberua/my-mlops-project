#!/usr/bin/env python3
"""
Trace hierarchical block connectivity from a connection name using DynamoDB GSI3.

Given a connection (e.g., v_8a2x9), traces the full signal chain across all
project scopes using targeted GSI3 (ConnectionIndex) queries:
  connection → Block (Input) → Block (Output) → next Block (Input) → ...

Each hop costs ~0.5 RCU instead of ~12,500 RCU for a full table scan at 100K items.

Usage:
    python trace_connection.py v_8a2x9 --profile myprofile
    python trace_connection.py v_8a2x9 --profile myprofile --format table
    python trace_connection.py v_8a2x9 --profile myprofile --format json
    python trace_connection.py v_8a2x9 --profile myprofile -o output.csv
"""

import argparse
import csv
import json
import sys


# ─── DynamoDB helpers ───

def _get_dynamodb_table(table_name, region, profile=None):
    """Create a boto3 DynamoDB Table resource.

    Helper that sets up a session with optional profile and returns
    a Table object ready for queries.

    Args:
        table_name: DynamoDB table name.
        region: AWS region.
        profile: Optional AWS CLI profile name.

    Returns:
        boto3 DynamoDB Table resource.
    """
    import boto3
    session_kwargs = {'region_name': region}
    if profile:
        session_kwargs['profile_name'] = profile
    session = boto3.Session(**session_kwargs)
    dynamodb = session.resource('dynamodb')
    return dynamodb.Table(table_name)


def _query_gsi3(table, connection, usage_prefix=None):
    """Query the ConnectionIndex (GSI3) for pins matching a connection name.

    GSI3 key structure:
        GSI3PK = CONN#{connection}
        GSI3SK = USE#{usage}|BLK#{block}|PIN#{pin}

    Optionally filters by usage prefix (e.g. 'Input' or 'Output') to narrow
    results to only input or output pins.

    Handles pagination automatically for large result sets.

    Args:
        table: boto3 DynamoDB Table resource.
        connection: Variable/connection name (e.g. 'v_8a2x9').
        usage_prefix: Optional — 'Input' or 'Output' to filter by pin usage.

    Returns:
        List of matching DynamoDB items.
    """
    from boto3.dynamodb.conditions import Key
    kce = Key('GSI3PK').eq(f'CONN#{connection}')
    if usage_prefix:
        kce = kce & Key('GSI3SK').begins_with(f'USE#{usage_prefix}')

    items = []
    response = table.query(IndexName='ConnectionIndex', KeyConditionExpression=kce)
    items.extend(response['Items'])
    while response.get('LastEvaluatedKey'):
        response = table.query(
            IndexName='ConnectionIndex',
            KeyConditionExpression=kce,
            ExclusiveStartKey=response['LastEvaluatedKey'],
        )
        items.extend(response['Items'])
    return items


def _query_block_pins(table, pk, block_execution, block_name):
    """Get ALL pins for a specific block by querying the base table.

    Uses the base table's PK + SK prefix to retrieve every pin (Input, Output,
    Const) belonging to a single block instance. This gives us the complete
    picture of a block — not just the pin that matched the connection.

    Args:
        table: boto3 DynamoDB Table resource.
        pk: Partition key (project scope), e.g. 'SYS#GREENLAND POWER|DEV#T1|...'
        block_execution: Execution order number of the block.
        block_name: Block name (e.g. 'MOVE_1', 'ADD_5').

    Returns:
        List of DynamoDB items — all pins for that block.
    """
    from boto3.dynamodb.conditions import Key
    prefix = f'BEX#{int(block_execution):04d}|BLK#{block_name}'
    response = table.query(
        KeyConditionExpression=Key('PK').eq(pk) & Key('SK').begins_with(prefix)
    )
    return response['Items']


# ─── Core trace function ───

def trace_chain_gsi3(start_connection, table, max_depth=10):
    """Trace the full signal chain using GSI3 queries (BFS per scope).

    This is the primary trace function. Given a starting variable name, it:
    1. Queries GSI3 to find all Input pins consuming that variable (SEED)
    2. Groups results by PK (project scope) — each scope traced independently
    3. For each scope, runs BFS:
       - Fetches all pins for each discovered block (base table query)
       - Follows Output connections into the BFS queue
       - Stops when queue is empty, max depth hit, or cycle detected
    4. Returns a dict of {scope_pk: [chain_items]}

    This avoids scanning the entire table. A typical 3-block trace uses
    ~7 queries (~3.5 RCU) vs ~12,500 RCU for a full scan at 100K items.

    Args:
        start_connection: Variable name to trace (e.g. 'v_8a2x9').
        table: boto3 DynamoDB Table resource.
        max_depth: Maximum BFS depth (default 10). Prevents runaway traces.

    Returns:
        Dict of {scope_pk: list_of_chain_items}. Each chain item is a dict
        with keys: depth, connection, block, block_type, block_execution,
        pin, pin_description, usage, data_type, project_context.
    """
    results = {}

    # Step 1: SEED — find where start_connection is consumed as Input
    seed_items = _query_gsi3(table, start_connection, usage_prefix='Input')
    if not seed_items:
        return results

    # Step 2: Group by project scope (PK)
    scopes = {}
    for item in seed_items:
        scopes.setdefault(item['PK'], []).append(item)

    # Step 3: Run BFS per scope
    for scope_pk, scope_seeds in scopes.items():
        visited_connections = {start_connection}
        visited_blocks = set()
        chain = []

        queue = [(start_connection, 0)]

        while queue:
            connection, depth = queue.pop(0)
            if depth > max_depth:
                continue

            # Find blocks consuming this connection as Input within this scope
            input_items = _query_gsi3(table, connection, usage_prefix='Input')
            input_items = [i for i in input_items if i['PK'] == scope_pk]

            for item in input_items:
                block = item['Block']
                bex = item['BlockExecution']
                block_key = (scope_pk, bex, block)

                if block_key in visited_blocks:
                    continue
                visited_blocks.add(block_key)

                # Get ALL pins for this block
                block_pins = _query_block_pins(table, scope_pk, bex, block)

                for pin in block_pins:
                    chain.append({
                        'depth': depth,
                        'connection': pin.get('Connection', ''),
                        'block': pin['Block'],
                        'block_type': pin.get('BlockType', ''),
                        'block_execution': str(pin.get('BlockExecution', '')),
                        'pin': pin['Pin'],
                        'pin_description': pin.get('PinDescription', ''),
                        'usage': pin['Usage'],
                        'data_type': pin.get('DataType', ''),
                        'project_context': scope_pk,
                    })

                    # Follow Output connections to the next block
                    if pin['Usage'] == 'Output':
                        out_conn = pin.get('Connection', '').strip()
                        if out_conn and out_conn not in visited_connections:
                            visited_connections.add(out_conn)
                            queue.append((out_conn, depth + 1))

        results[scope_pk] = chain

    # Sort each scope by depth then block execution order
    for scope_pk in results:
        results[scope_pk].sort(
            key=lambda x: (x['depth'], int(x['block_execution']) if x['block_execution'].isdigit() else 0)
        )
    return results


# ─── Output formatters ───

def print_chain(chain, start_connection, all_rows):
    """Print the trace result as a visual tree with full block details.

    Two sections:
    1. Connection Flow — indented tree showing signal path with arrows
       (⬤ for depth 0, → for deeper hops), pin direction (◀ IN / ▶ OUT),
       and data types.
    2. Full Block Details — tabular dump of every column for each block
       in the chain, useful for debugging and verification.

    Args:
        chain: List of chain item dicts from trace_chain_gsi3().
        start_connection: The variable name that was traced.
        all_rows: All pin rows (used for the full block details section).
    """
    if not chain:
        print(f"No connections found for: {start_connection}")
        return

    contexts = list(dict.fromkeys(item.get('project_context', '') for item in chain))
    print(f"\n{'='*80}")
    print(f"CONNECTION FLOW: {start_connection}")
    for ctx in contexts:
        if ctx:
            print(f"PROJECT: {ctx}")
    print(f"{'='*80}\n")

    prev_block = None
    for item in chain:
        indent = "  " * item['depth']
        block = item['block']
        pin = item['pin']
        usage = item['usage']
        conn = item['connection']
        block_type = item['block_type']
        bex = item['block_execution']
        data_type = item['data_type']

        if block != prev_block:
            arrow = "→" if item['depth'] > 0 else "⬤"
            print(f"{indent}{arrow} [{block_type}] {block} (Execution: {bex})")
            prev_block = block

        direction = "◀ IN " if usage == 'Input' else "▶ OUT"
        print(f"{indent}    {direction}  {pin:<10} = {conn:<20} ({data_type})")

    blocks = list(dict.fromkeys(item['block'] for item in chain))
    print(f"\nBlocks traversed: {' → '.join(blocks)}")

    # Full Row Details Per Block
    print(f"\n{'='*80}")
    print(f"FULL BLOCK DETAILS")
    print(f"{'='*80}")

    block_rows = {}
    for row in all_rows:
        b = row.get('Block', '').strip()
        block_rows.setdefault(b, []).append(row)

    if not all_rows:
        return
    columns = list(all_rows[0].keys())

    for block_name in blocks:
        rows = block_rows.get(block_name, [])
        if not rows:
            continue

        print(f"\n{'─'*80}")
        print(f"  Block: {block_name} ({len(rows)} pins)")
        print(f"{'─'*80}")

        col_widths = {}
        for col in columns:
            max_val = max((len(str(row.get(col, ''))) for row in rows), default=0)
            col_widths[col] = max(len(col), min(max_val, 40))

        display_cols = [c for c in columns if c != 'Locator']
        header = " | ".join(f"{col:<{col_widths[col]}}" for col in display_cols)
        print(f"  {header}")
        print(f"  {'-' * len(header)}")

        for row in rows:
            line = " | ".join(
                f"{str(row.get(col, '')):<{col_widths[col]}}"
                for col in display_cols
            )
            print(f"  {line}")

    print(f"\n{'='*80}")
    print(f"Total blocks: {len(blocks)} | Total pins in chain: {len(chain)}")
    print(f"{'='*80}\n")


def print_chain_table(chain, start_connection):
    """Print the trace result as a flat formatted table.

    Compact output with columns: Depth, Block, Type, Exec, Pin, Usage,
    Connection, DataType. One row per pin. Useful for quick inspection
    and copy-pasting into docs.

    Args:
        chain: List of chain item dicts.
        start_connection: The variable name that was traced.
    """
    if not chain:
        print(f"No connections found for: {start_connection}")
        return

    print(f"\nConnection Trace: {start_connection}\n")

    header = f"{'Depth':<6} {'Block':<12} {'Type':<10} {'Exec':<5} {'Pin':<10} {'Usage':<7} {'Connection':<20} {'DataType':<8}"
    print(header)
    print("-" * len(header))

    for item in chain:
        print(
            f"{item['depth']:<6} "
            f"{item['block']:<12} "
            f"{item['block_type']:<10} "
            f"{item['block_execution']:<5} "
            f"{item['pin']:<10} "
            f"{item['usage']:<7} "
            f"{item['connection']:<20} "
            f"{item['data_type']:<8}"
        )


# ─── Export functions ───

def export_results(chain, all_rows, start_connection, output_path):
    """Export trace results to a CSV or Excel file.

    Adds metadata columns (ChainOrder, ChainDepth, TracedConnection) to each
    row so the export is self-describing. For Excel, creates two sheets:
    'Connection Flow' (summary) and 'Block Details' (full rows).

    Args:
        chain: List of chain item dicts.
        all_rows: All pin rows for blocks in the chain.
        start_connection: The variable name that was traced.
        output_path: File path — .csv or .xlsx.
    """
    blocks = list(dict.fromkeys(item['block'] for item in chain))

    block_rows = {}
    for row in all_rows:
        b = row.get('Block', '').strip()
        if b in blocks:
            block_rows.setdefault(b, []).append(row)

    columns = list(all_rows[0].keys()) if all_rows else []

    export_rows = []
    block_depth = {}
    for item in chain:
        if item['block'] not in block_depth:
            block_depth[item['block']] = item['depth']

    for order, block_name in enumerate(blocks, 1):
        for row in block_rows.get(block_name, []):
            out_row = {
                'ChainOrder': order,
                'ChainDepth': block_depth.get(block_name, 0),
                'TracedConnection': start_connection,
            }
            for col in columns:
                out_row[col] = row.get(col, '')
            export_rows.append(out_row)

    export_columns = ['ChainOrder', 'ChainDepth', 'TracedConnection'] + columns

    if output_path.endswith(('.xlsx', '.xls')):
        _export_excel(export_rows, export_columns, output_path, start_connection, chain, blocks)
    else:
        _export_csv(export_rows, export_columns, output_path)

    print(f"✓ Exported {len(export_rows)} rows to {output_path}")


def _export_csv(export_rows, columns, output_path):
    """Write trace results to a CSV file.

    Args:
        export_rows: List of dicts to write.
        columns: Ordered list of column names.
        output_path: Destination file path.
    """
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=columns)
        writer.writeheader()
        writer.writerows(export_rows)


def _export_excel(export_rows, columns, output_path, start_connection, chain, blocks):
    """Write trace results to an Excel file with two sheets.

    Sheet 1 — 'Connection Flow': Summary table of the BFS trace
    (Depth, Block, BlockType, Execution, Pin, Usage, Connection, DataType).

    Sheet 2 — 'Block Details': Full row data for every pin in every
    block that appeared in the chain, with all original CSV columns.

    Both sheets get styled headers and auto-fitted column widths.

    Args:
        export_rows: List of dicts for the detail sheet.
        columns: Ordered column names for the detail sheet.
        output_path: Destination .xlsx file path.
        start_connection: The traced variable name.
        chain: List of chain item dicts for the flow sheet.
        blocks: Ordered list of block names in chain order.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print("openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.Workbook()

    # Sheet 1: Connection Flow
    ws_flow = wb.active
    ws_flow.title = "Connection Flow"
    flow_headers = ['Depth', 'Block', 'BlockType', 'Execution', 'Pin', 'Usage', 'Connection', 'DataType']
    for col_idx, h in enumerate(flow_headers, 1):
        cell = ws_flow.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for row_idx, item in enumerate(chain, 2):
        ws_flow.cell(row=row_idx, column=1, value=item['depth'])
        ws_flow.cell(row=row_idx, column=2, value=item['block'])
        ws_flow.cell(row=row_idx, column=3, value=item['block_type'])
        ws_flow.cell(row=row_idx, column=4, value=item['block_execution'])
        ws_flow.cell(row=row_idx, column=5, value=item['pin'])
        ws_flow.cell(row=row_idx, column=6, value=item['usage'])
        ws_flow.cell(row=row_idx, column=7, value=item['connection'])
        ws_flow.cell(row=row_idx, column=8, value=item['data_type'])

    for col in ws_flow.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws_flow.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    # Sheet 2: Full Block Details
    ws_detail = wb.create_sheet("Block Details")
    for col_idx, h in enumerate(columns, 1):
        cell = ws_detail.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for row_idx, row_data in enumerate(export_rows, 2):
        for col_idx, col_name in enumerate(columns, 1):
            ws_detail.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ''))

    for col in ws_detail.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws_detail.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    wb.save(output_path)


# ─── CLI entry point ───

def main():
    """CLI entry point — parses args and runs the GSI3-based trace.

    Connects to DynamoDB, traces the given variable across all project scopes
    using GSI3 queries, and outputs results in the chosen format (tree, table,
    or JSON). Also exports a CSV per scope automatically.
    """
    parser = argparse.ArgumentParser(
        description='Trace hierarchical block connectivity from a connection name using DynamoDB GSI3'
    )
    parser.add_argument('connection', help='Connection to trace (e.g., v_8a2x9)')
    parser.add_argument('--table', default='markvie-kb-poc', help='DynamoDB table name')
    parser.add_argument('--region', default='us-west-2', help='AWS region')
    parser.add_argument('--profile', default=None, help='AWS profile')
    parser.add_argument('--max-depth', type=int, default=10, help='Max BFS depth (default: 10)')
    parser.add_argument('--format', choices=['tree', 'table', 'json'], default='tree',
                        help='Output format (default: tree)')
    parser.add_argument('--output', '-o', default=None,
                        help='Save results to file (.csv or .xlsx)')
    args = parser.parse_args()

    print(f"Tracing {args.connection} via GSI3 on {args.table}")

    # Connect to DynamoDB and trace using GSI3 queries
    ddb_table = _get_dynamodb_table(args.table, args.region, args.profile)
    results = trace_chain_gsi3(args.connection, ddb_table, max_depth=args.max_depth)

    total_pins = sum(len(chain) for chain in results.values())
    total_blocks = sum(len(set(i['block'] for i in chain)) for chain in results.values())
    print(f"Trace complete: {len(results)} scope(s), {total_pins} pins across {total_blocks} blocks")

    # Output each scope separately
    for scope_pk, chain in results.items():
        print(f"\n{'='*80}")
        print(f"SCOPE: {scope_pk}")
        print(f"{'='*80}")

        # Build row dicts for display/export from chain items
        all_rows = []
        for item in chain:
            all_rows.append({
                'Pin': item['pin'],
                'PinDescription': item['pin_description'],
                'Block': item['block'],
                'BlockDescription': '',
                'Task': '',
                'Program': '',
                'Locator': '',
                'BlockExecution': item['block_execution'],
                'BlockType': item['block_type'],
                'Connection': item['connection'],
                'DataType': item['data_type'],
                'EntryNo': '0',
                'IsCritical': 'FALSE',
                'ProgramExecution': '0',
                'Usage': item['usage'],
            })

        if args.format == 'json':
            blocks = list(dict.fromkeys(item['block'] for item in chain))
            block_rows = {}
            for row in all_rows:
                b = row.get('Block', '').strip()
                if b in blocks:
                    block_rows.setdefault(b, []).append(row)
            output = {
                'connection': args.connection,
                'scope': scope_pk,
                'flow': chain,
                'blocks': block_rows,
            }
            print(json.dumps(output, indent=2))
        elif args.format == 'table':
            print_chain_table(chain, args.connection)
        else:
            print_chain(chain, args.connection, all_rows)

        # Export per scope
        safe_scope = scope_pk.replace('#', '').replace('|', '_')[:40]
        output_path = args.output if args.output else f"trace_{args.connection}_{safe_scope}.csv"
        export_results(chain, all_rows, args.connection, output_path)


if __name__ == '__main__':
    main()
